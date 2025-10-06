import re
import json
import io
import unicodedata
from typing import Tuple, Optional, Dict, List, Set
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# ======================= Look & feel "app" =======================
st.set_page_config(
    page_title="Gestione Clienti",
    page_icon="icon-512.png",      # cambia in "static/icon-512.png" se usi /static
    layout="wide"
)
st.markdown("""
<link rel="apple-touch-icon" sizes="180x180" href="/apple-touch-icon.png">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-title" content="Gestione Clienti">
<style>
#MainMenu, header, footer {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

# =========================== Utility ============================
def normalize_text(s: str) -> str:
    if s is None:
        return ""
    if not isinstance(s, str):
        s = str(s)
    s = unicodedata.normalize("NFKD", s).encode("ASCII", "ignore").decode("ASCII")
    s = s.lower()
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def get_first_nonempty(values) -> str:
    for v in values:
        if pd.notna(v):
            vs = str(v).strip()
            if vs and vs.lower() != "none":
                return vs
    return ""

def clean_table(df: pd.DataFrame) -> pd.DataFrame:
    """Rimuove colonne/righe completamente vuote o 'None', sostituisce 'None' con ''."""
    if df is None or df.empty:
        return pd.DataFrame()
    tmp = df.copy()
    tmp = tmp.astype(object).where(~tmp.isna(), None)
    tmp = tmp.applymap(lambda x: "" if (x is None or str(x).strip().lower() == "none") else x)
    # drop colonne vuote
    tmp = tmp.replace("", pd.NA).dropna(axis=1, how="all").fillna("")
    # drop righe vuote
    tmp = tmp.replace("", pd.NA).dropna(axis=0, how="all").fillna("")
    return tmp

# ======= Parser: Contratti di Noleggio + Note (pandas) =======
def parse_contracts_and_notes(sheet_df: pd.DataFrame) -> Tuple[pd.DataFrame, str, int]:
    """
    Ritorna (contratti_df, note_text, header_row_openpyxl_hint)
    - header_row_openpyxl_hint: indice 0-based della riga header in pandas (serve per allineamento con openpyxl).
    """
    df = sheet_df.copy()
    df = df.dropna(axis=1, how="all")
    df = df.astype(str).where(~df.isna(), None)

    col0 = df.columns[0]
    first_col = df[col0].apply(lambda x: str(x).strip() if x is not None else "")

    header_row = None
    for idx, key in first_col.items():
        if normalize_text(key).startswith("contratti di noleggio"):
            header_row = idx + 1  # riga header (con le etichette)
            break

    contratti_df = pd.DataFrame()
    if header_row is not None and header_row < len(df):
        headers = [str(x).strip() if pd.notna(x) else "" for x in df.iloc[header_row].tolist()]
        headers = [h if h else f"Col_{i}" for i, h in enumerate(headers)]
        rows = []
        for r in range(header_row + 1, len(df)):
            row0 = str(df.iloc[r, 0]).strip() if pd.notna(df.iloc[r, 0]) else ""
            if normalize_text(row0).startswith("note clienti"):
                break
            # riga non tutta vuota?
            if not all((str(x).strip() == "" or str(x).strip().lower() == "none") for x in df.iloc[r].tolist()):
                rows.append([None if str(x).strip().lower() == "none" else x for x in df.iloc[r].tolist()])
        if rows:
            contratti_df = pd.DataFrame(rows, columns=headers)
            # parse date (nomi colonne contenenti "data")
            for c in list(contratti_df.columns):
                if "data" in normalize_text(c):
                    contratti_df[c] = pd.to_datetime(contratti_df[c], errors="coerce", dayfirst=True)

    # NOTE CLIENTI
    note_text = ""
    for idx, key in first_col.items():
        if normalize_text(key).startswith("note clienti"):
            rr = idx + 1
            if rr < len(df):
                note_text = get_first_nonempty([df.at[rr, c] for c in df.columns])
            break

    # pulizia + format date → dd/mm/yy
    if not contratti_df.empty:
        for c in contratti_df.columns:
            if pd.api.types.is_datetime64_any_dtype(contratti_df[c]):
                contratti_df[c] = contratti_df[c].dt.strftime("%d/%m/%y")
        contratti_df = clean_table(contratti_df)

    return contratti_df, note_text, (header_row if header_row is not None else -1)

# ---------- Parser: Dati cliente (chiave/valore prima dei contratti) ----------
def parse_client_info(sheet_df: pd.DataFrame) -> Tuple[str, Dict[str, str]]:
    df = sheet_df.copy()
    df = df.dropna(axis=1, how="all")
    df = df.astype(str).where(~df.isna(), None)

    col0 = df.columns[0]
    first_col = df[col0].apply(lambda x: str(x).strip() if x is not None else "")

    # righe di stop
    stop_rows = []
    for idx, key in first_col.items():
        nk = normalize_text(key)
        if nk.startswith("contratti di noleggio") or nk.startswith("note clienti"):
            stop_rows.append(idx)
    stop_at = min(stop_rows) if stop_rows else len(df)

    # nome cliente
    nome = ""
    for idx, key in first_col.items():
        if idx >= stop_at:
            break
        nk = normalize_text(key)
        if nk in ("nome cliente", "cliente"):
            nome = get_first_nonempty([df.at[idx, c] for c in df.columns[1:]])
            if nome:
                break

    # chiave -> valore
    info: Dict[str, str] = {}
    SKIP = {"scheda cliente", "torna all indice", "totale contratti", "dati cliente", "cliente", "nome cliente"}
    for idx, key in first_col.items():
        if idx >= stop_at:
            break
        k_raw = str(key).strip()
        if not k_raw:
            continue
        if normalize_text(k_raw) in SKIP:
            continue
        v = get_first_nonempty([df.at[idx, c] for c in df.columns[1:]])
        if v:
            vv = "" if str(v).strip().lower() == "none" else str(v).strip()
            if vv:
                info[k_raw] = vv

    return nome, info

# ===== Indice: elenco dall'Indice + filtri per nome e città =====
def extract_client_list_from_indice(indice_df: pd.DataFrame) -> pd.DataFrame:
    """Ritorna DataFrame con colonne: Nome, Città (se presente), Telefono (se presente)."""
    if indice_df is None or indice_df.empty:
        return pd.DataFrame(columns=["Nome", "Città", "Telefono"])

    header_row0 = indice_df.iloc[0].to_dict()
    col_cli = None
    for c, v in header_row0.items():
        if isinstance(v, str) and "cliente" in v.lower():
            col_cli = c
            break
    if col_cli is None:
        candidates = [c for c in indice_df.columns if indice_df[c].notna().any()]
        col_cli = candidates[0] if candidates else indice_df.columns[0]

    col_citta = None
    col_tel = None
    for c in indice_df.columns:
        v0 = str(indice_df.at[0, c]) if 0 in indice_df.index else ""
        nk = normalize_text(v0)
        if not col_citta and ("citta" in nk or "citt" in nk): col_citta = c
        if not col_tel and ("telefono" in nk or "tel" in nk): col_tel = c

    names = indice_df[col_cli].iloc[1:].dropna().astype(str).map(str.strip).tolist() if col_cli in indice_df.columns else []
    cities = indice_df[col_citta].iloc[1:].astype(str).map(str.strip).tolist() if col_citta and col_citta in indice_df.columns else []
    tels = indice_df[col_tel].iloc[1:].astype(str).map(str.strip).tolist() if col_tel and col_tel in indice_df.columns else []

    maxlen = max(len(names), len(cities), len(tels), 0)
    def safe(lst, i): return lst[i] if i < len(lst) else ""
    rows, seen = [], set()
    for i in range(maxlen):
        nome = safe(names, i)
        if not nome or normalize_text(nome) in ("", "cliente"): continue
        key = normalize_text(nome)
        if key in seen: continue
        seen.add(key)
        rows.append({"Nome": nome, "Città": safe(cities, i), "Telefono": safe(tels, i)})
    out = pd.DataFrame(rows, columns=["Nome", "Città", "Telefono"]).fillna("")
    return out

def find_client_sheet_name(sheets: Dict[str, pd.DataFrame], cliente: str) -> Optional[str]:
    target = normalize_text(cliente)
    for name in sheets.keys():
        if normalize_text(name) == target: return name
    for name in sheets.keys():
        if normalize_text(name).startswith(target): return name
    for name in sheets.keys():
        if target in normalize_text(name): return name
    return None

# ===== Rileva righe “rosse” tramite openpyxl =====
def detect_red_rows(uploaded_bytes: bytes, sheet_name: str) -> Tuple[Set[int], List[str]]:
    """
    Restituisce:
      - set di indici 0-based delle righe dati dei contratti con evidenza rossa,
      - intestazioni lette dalla riga header.
    Logica:
      - cerca "Contratti di Noleggio" nella prima colonna,
      - header = riga successiva,
      - dati = righe successive fino a "NOTE CLIENTI" o riga vuota,
      - una riga è "rossa" se almeno una cella ha fill rosso (rgb che finisce con 'FF0000').
    """
    buf = io.BytesIO(uploaded_bytes)
    wb = load_workbook(buf, data_only=True)
    if sheet_name not in wb.sheetnames:
        return set(), []
    ws = wb[sheet_name]

    def cell_text(r, c) -> str:
        v = ws.cell(row=r, column=c).value
        return str(v).strip() if v is not None else ""

    # trova riga titolo
    title_row = None
    for r in range(1, ws.max_row + 1):
        val = cell_text(r, 1)
        if normalize_text(val).startswith("contratti di noleggio"):
            title_row = r
            break
    if title_row is None:
        return set(), []

    header_row = title_row + 1
    headers = [cell_text(header_row, c) for c in range(1, ws.max_column + 1)]
    headers = [h if h else f"Col_{i-1}" for i, h in enumerate(headers, start=1)]

    red_rows: Set[int] = set()
    data_index = 0  # indice 0-based sui dati (allineato al DataFrame dei contratti)
    r = header_row + 1
    while r <= ws.max_row:
        first = cell_text(r, 1)
        if normalize_text(first).startswith("note clienti"):
            break
        # riga completamente vuota?
        row_values = [cell_text(r, c) for c in range(1, ws.max_column + 1)]
        if all(v == "" for v in row_values):
            # salta vuote, ma NON interrompe (alcuni file hanno righe vuote intermedie)
            r += 1
            continue

        # check fill rosso
        is_red = False
        for c in range(1, ws.max_column + 1):
            fill = ws.cell(row=r, column=c).fill
            if fill and getattr(fill, "patternType", None) == "solid":
                col = getattr(fill, "fgColor", None)
                rgb = getattr(col, "rgb", None)
                if rgb and str(rgb).upper().endswith("FF0000"):
                    is_red = True
                    break
        if is_red:
            red_rows.add(data_index)

        data_index += 1
        r += 1

    return red_rows, headers

def style_html_table(df: pd.DataFrame, red_idx: Set[int]) -> str:
    """Rende una tabella HTML con background rosso chiaro sulle righe marcate."""
    # escape semplice
    def esc(x): 
        s = "" if x is None else str(x)
        return s.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
    # header
    html = ['<div style="overflow:auto;"><table style="border-collapse:collapse;width:100%;">']
    html.append("<thead><tr>")
    for c in df.columns:
        html.append(f'<th style="text-align:left;border-bottom:1px solid #ddd;padding:6px;">{esc(c)}</th>')
    html.append("</tr></thead><tbody>")
    # rows
    for i, (_, row) in enumerate(df.iterrows()):
        bg = "background-color:#ffe5e5;" if i in red_idx else ""
        html.append(f'<tr style="{bg}">')
        for v in row.tolist():
            html.append(f'<td style="padding:6px;border-bottom:1px solid #f0f0f0;">{esc(v)}</td>')
        html.append("</tr>")
    html.append("</tbody></table></div>")
    return "".join(html)

# ============================ STATE ============================
if "view" not in st.session_state:
    st.session_state.view = "Indice / Anagrafiche"   # oppure "Scheda Cliente"
if "selected_cliente" not in st.session_state:
    st.session_state.selected_cliente = None
if "notes_store" not in st.session_state:
    st.session_state.notes_store = {}  # {cliente -> nota}
if "info_overrides" not in st.session_state:
    st.session_state.info_overrides = {}  # {cliente -> dict info editata}

# ============================ FILE UPLOAD ============================
uploaded = st.file_uploader("📥 Carica il file Excel (.xlsx/.xlsm)", type=["xlsx", "xlsm"])
if not uploaded:
    st.info("Carica il file per iniziare.")
    st.stop()

uploaded_bytes = uploaded.getvalue()  # serve per openpyxl

@st.cache_data(show_spinner=False)
def load_all_sheets(file) -> Dict[str, pd.DataFrame]:
    return pd.read_excel(file, sheet_name=None, engine="openpyxl", dtype=str)

sheets_dict = load_all_sheets(uploaded)
if not sheets_dict:
    st.error("Nessun foglio trovato nel file.")
    st.stop()

# Trova "Indice"
names_map = {n.lower(): n for n in sheets_dict.keys()}
indice_key = names_map.get("indice")
indice_df = sheets_dict[indice_key] if indice_key else pd.DataFrame()
index_table = extract_client_list_from_indice(indice_df)

# ============================ NAV ============================
st.sidebar.title("Navigazione")
page = st.sidebar.radio(
    "Vai a…",
    options=["Indice / Anagrafiche", "Scheda Cliente"],
    index=0 if st.session_state.view == "Indice / Anagrafiche" else 1
)
st.session_state.view = page

# ============================ PAGINA: INDICE ============================
if st.session_state.view == "Indice / Anagrafiche":
    st.title("📇 Indice Clienti")

    colf1, colf2, colf3 = st.columns([2,2,1])
    with colf1:
        q_name = st.text_input("🔎 Cerca per nome", value="", placeholder="es. Rossi, 2 ESSE…")
    with colf2:
        q_city = st.text_input("🏙️ Cerca per città", value="", placeholder="es. Milano, Casarile…")
    with colf3:
        st.write(""); st.write("")
        if st.button("🔄 Pulisci filtri"):
            st.rerun()

    filt = index_table.copy()
    if q_name:
        filt = filt[filt["Nome"].astype(str).map(lambda x: normalize_text(q_name) in normalize_text(x))]
    if q_city:
        filt = filt[filt["Città"].astype(str).map(lambda x: normalize_text(q_city) in normalize_text(x))]

    st.caption(f"{len(filt)} clienti trovati")
    st.dataframe(filt, use_container_width=True, hide_index=True)

    choices: List[str] = ["-- Seleziona --"] + filt["Nome"].tolist()
    pick = st.selectbox("Apri scheda cliente", choices)
    if pick and pick != "-- Seleziona --":
        st.session_state.selected_cliente = pick
        st.session_state.view = "Scheda Cliente"
        st.rerun()

# ============================ PAGINA: SCHEDA ============================
if st.session_state.view == "Scheda Cliente":
    st.title("🧾 Scheda Cliente")
    st.button("← Torna all’Indice", on_click=lambda: (st.session_state.update({"view":"Indice / Anagrafiche","selected_cliente":None}), st.rerun()))

    cliente_sel: Optional[str] = st.session_state.selected_cliente
    if not cliente_sel:
        # permetti scelta rapida se si entra direttamente
        choices: List[str] = ["-- Seleziona --"] + index_table["Nome"].tolist()
        pick = st.selectbox("Seleziona cliente", choices)
        if pick and pick != "-- Seleziona --":
            st.session_state.selected_cliente = pick
            st.rerun()
        st.stop()

    foglio = find_client_sheet_name(sheets_dict, cliente_sel)
    if not foglio:
        st.warning("Foglio cliente non trovato.")
        st.stop()

    sheet_df = sheets_dict[foglio]
    nome_cli, info_cli = parse_client_info(sheet_df)
    contratti_df, note_esistente, header_hint = parse_contracts_and_notes(sheet_df)
    note_val = st.session_state.notes_store.get(cliente_sel, note_esistente or "")

    # Applica override dei dati cliente (modifiche da maschera)
    if cliente_sel in st.session_state.info_overrides:
        info_cli = {**info_cli, **st.session_state.info_overrides[cliente_sel]}

    st.markdown(f"## {cliente_sel}")
    if nome_cli and normalize_text(nome_cli) != normalize_text(cliente_sel):
        st.caption(f"Nome da scheda: {nome_cli}")

    # ----- MASCHERA DATI CLIENTE (solo anagrafica) -----
    with st.expander("✏️ Maschera Dati Cliente", expanded=False):
        base_order = ["Indirizzo", "Città", "CAP", "TELEFONO", "MAIL", "RIF.", "RIF 2.", "IBAN", "partita iva", "SDI"]
        extra = [k for k in info_cli.keys() if k not in base_order]
        keys = base_order + extra

        with st.form("form_info_cliente"):
            c1, c2 = st.columns(2)
            edited = {}
            for i, k in enumerate(keys):
                target = c1 if i % 2 == 0 else c2
                with target:
                    edited[k] = st.text_input(k, value=info_cli.get(k, ""))
            saved = st.form_submit_button("💾 Salva Dati Cliente (sessione)")
        if saved:
            st.session_state.info_overrides[cliente_sel] = edited
            info_cli = {**info_cli, **edited}
            st.success("Dati cliente aggiornati (solo in questa sessione).")

    # ----- DATI CLIENTE SOPRA -----
    st.subheader("👤 Dati Cliente")
    if info_cli:
        ordered = ["Indirizzo", "Città", "CAP", "TELEFONO", "MAIL", "RIF.", "RIF 2.", "IBAN", "partita iva", "SDI", "Ultimo Recall", "ultima visita"]
        keys = [k for k in ordered if k in info_cli] + [k for k in info_cli.keys() if k not in ordered]
        for k in keys:
            st.markdown(f"**{k}**")
            st.write(info_cli[k])
    else:
        st.caption("Nessun dato anagrafico trovato.")

    # ----- CONTRATTI SOTTO (con evidenza rosse) -----
    st.subheader("📑 Contratti di Noleggio")

    red_idx, headers_from_xl = detect_red_rows(uploaded_bytes, foglio)

    if contratti_df is not None and not contratti_df.empty:
        # Aggiungi colonna Evidenza
        pretty = clean_table(contratti_df)
        # riallinea eventuali indici (red_idx è 0-based sui dati)
        flags = ["🔴" if i in red_idx else "" for i in range(len(pretty))]
        pretty = pretty.copy()
        pretty.insert(0, "Evidenza", flags)

        st.checkbox("Colora righe rosse", value=False, key="colorize")
        if st.session_state.colorize and len(pretty) > 0:
            # tabella HTML con righe colorate
            html = style_html_table(pretty, {i for i in range(len(flags)) if flags[i] == "🔴"})
            st.markdown(html, unsafe_allow_html=True)
        else:
            st.dataframe(pretty, use_container_width=True, hide_index=True)
    else:
        st.info("Nessun contratto trovato in questa scheda.")

    # ----- NOTE -----
    st.subheader("📝 Note Cliente")
    new_note = st.text_area("Testo note", value=note_val, height=140, placeholder="Scrivi o aggiorna le note qui…")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("💾 Salva nota (solo in questa sessione)"):
            st.session_state.notes_store[cliente_sel] = new_note
            st.success("Nota salvata nella sessione corrente.")
    with c2:
        notes_json = json.dumps(st.session_state.notes_store, ensure_ascii=False, indent=2)
        st.download_button("⬇️ Esporta tutte le note (JSON)", data=notes_json, file_name="note_clienti.json", mime="application/json")
